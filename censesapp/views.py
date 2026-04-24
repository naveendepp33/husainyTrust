from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth.models import User
from django.contrib import messages
from .models import *
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from datetime import date, datetime
from django.db.models import Q
from django.http import HttpResponse
import openpyxl
from reportlab.pdfgen import canvas

# Create your views here.

def base(request):
    news = News.objects.first()
    print(news)
    return render(request, "base.html", {
        "site_news": news
    })
   
def userlogout(request):
    logout(request)
    return redirect("login")   


def logins(request):
    if request.method == "POST":
        login_input = request.POST.get("login_input", "").strip()
        password = request.POST.get("password")
        remember_me = request.POST.get("remember_me")

        user = None

        # 🔹 Step 1: Try Employee ID
        profile = AddUser.objects.filter(employee_id__iexact=login_input).first()
        if profile:
            user = authenticate(request, username=profile.user.username, password=password)

        # 🔹 Step 2: If not found, try Username
        if user is None:
            user = authenticate(request, username=login_input, password=password)

        # 🔹 Step 3: Validate User
        if user is not None:
            profile = AddUser.objects.filter(user=user).first()

            # Superuser case
            if not profile:
                login(request, user)
                return redirect("dashboard")

            # Volunteer check
            if profile.role == "volunteer":
                today = date.today()
                now_time = datetime.now().time()

                # 🔹 Step 1: Find allocations where today is between start_date and end_date
                allocations = VolunteerAllocation.objects.filter(
                    volunteer=profile,
                    start_date__lte=today,
                    end_date__gte=today
                )

                # 🔹 Step 2: Check if current time is within the shift of ANY of those allocations
                allowed = False
                for a in allocations:
                    # If start/end time are not set, you might want to allow 24h access 
                    # or block it. Here we assume they must be set:
                    if a.start_time and a.end_time:
                        if a.start_time <= now_time <= a.end_time:
                            allowed = True
                            break
                
                if not allowed:
                    messages.error(request, "Access Denied: You are not allocated to work at this date or time.")
                    return redirect("/")

            if not remember_me:
                request.session.set_expiry(0)

            login(request, user)

            if profile.role == "volunteer":
                return redirect("/addfamily/")
            else:
                return redirect("dashboard")

        else:
            messages.error(request, "Invalid Employee ID / Username or Password")
            return redirect("/")

    return render(request, "login.html")

@login_required(login_url="/")
def dashboard(request):
    return render(request, 'dashboard.html')

@login_required(login_url="/")
def adduser(request):

    supervisors = AddUser.objects.filter(role="manager")

    role = request.GET.get("role")   
    next_employee_id = ""

    
    if role:
        prefix_map = {
            "admin": "ADM",
            "manager": "SVR",
            "volunteer": "VOL"
        }

        prefix = prefix_map.get(role)

        last_user = AddUser.objects.filter(role=role).order_by("-id").first()

        if last_user and last_user.employee_id:
            last_number = int(last_user.employee_id[-3:])
            new_number = last_number + 1
        else:
            new_number = 1

        next_employee_id = f"{prefix}{new_number:04d}"


    # 🔹 Handle POST Save
    if request.method == "POST":
        role = request.POST.get("role")
        # fullname = request.POST.get("fullname")
        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        username = (request.POST.get("username") or "").title()
        password = request.POST.get("password")
        supervisor_id = request.POST.get("supervisor")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists")
            return redirect("adduser")

        # generate again for safety
        prefix_map = {
            "admin": "ADM",
            "manager": "SVR",
            "volunteer": "VOL"
        }

        prefix = prefix_map.get(role)
        last_user = AddUser.objects.filter(role=role).order_by("-id").first()

        if last_user and last_user.employee_id:
            last_number = int(last_user.employee_id[-3:])
            new_number = last_number + 1
        else:
            new_number = 1

        employee_id = f"{prefix}{new_number:04d}"

        user = User(username=username, email=email)
        user.set_password(password)
        user.save()

        AddUser.objects.create(
            user=user,
            role=role,
            employee_id=employee_id,
            
            email=email,
            mobile=mobile,
            username=username,
            password=user.password,
            supervisor_id=supervisor_id
        )

        messages.success(request, "User Created Successfully")
        return redirect("manageuser")

    return render(request, "adduser.html", {
        "next_employee_id": next_employee_id,
        "selected_role": role,
        "supervisors": supervisors,
    })

@login_required(login_url="/")
def manageuser(request):
    supervisors = AddUser.objects.filter(role="manager")
    role_filter = request.GET.get("role")

    if role_filter:
        users = AddUser.objects.filter(role=role_filter)
    else:
        users = AddUser.objects.all()

    context = {
        "users": users,
        "supervisors": supervisors,
        "selected_role": role_filter
    }

    return render(request, 'manageuser.html', context)

@login_required(login_url="/")
def view_user_details(request, id):
    try:
        user = AddUser.objects.select_related('user').get(id=id)

        data = {
            "employee_id": user.employee_id,
            # "fullname": user.fullname,
            "email": user.user.email,
            "mobile": user.mobile,
            "username": user.user.username,
            "role": user.get_role_display(),
            
        }

        return JsonResponse(data)

    except AddUser.DoesNotExist:
        return JsonResponse({"error": "User not found"})

@login_required(login_url="/")  
def get_user_data(request, id):
    user = AddUser.objects.select_related('user').get(id=id)

    data = {
        "id": user.id,
        "employee_id": user.employee_id,
        "role": user.role,
        # "fullname": user.fullname,
        "email": user.user.email,
        "mobile": user.mobile,
        "username": user.user.username,
    }

    return JsonResponse(data)

@csrf_exempt
@login_required(login_url="/")
def update_user(request):
    if request.method == "POST":
        id = request.POST.get("id")
        user_profile = AddUser.objects.select_related('user').get(id=id)
        user = user_profile.user

        email = request.POST.get("email")
        mobile = request.POST.get("mobile")
        username = (request.POST.get("username") or "").title()
        password = request.POST.get("password")
        old_password = request.POST.get("old_password")

        if password:
            if not old_password:
                return JsonResponse({"success": False, "error": "Old password is required to change password."})
            
            if not user.check_password(old_password):
                return JsonResponse({"success": False, "error": "Incorrect old password."})
            
            user.set_password(password)

        user.username = username
        user.email = email
        user_profile.mobile = mobile

        user.save()
        user_profile.save()

        return JsonResponse({"success": True})

    return JsonResponse({"success": False, "error": "Invalid request method."})

@login_required(login_url="/")
def delete_user(request, user_id):
    try:
        user = AddUser.objects.get(id=user_id)
        user.user.delete()  # This will also delete the AddUser due to OneToOne relation
        messages.success(request, "User deleted successfully.")
    except AddUser.DoesNotExist:
        messages.error(request, "User does not exist.")
    
    return redirect("manageuser")

@login_required(login_url="/")
def rolemanagement(request):
    return render(request, 'rolemanagement.html')

@login_required(login_url="/")
def grade(request):
    grades = Grade.objects.all().order_by("-id")

    if request.method == "POST":
        grade_name = request.POST.get("grade_name")

        if Grade.objects.filter(name__iexact=grade_name).exists():
            messages.error(request, "Grade already exists!")
            return redirect("grade")
        
        Grade.objects.create(name=grade_name)

        messages.success(request, "Grade Added Successfully!")
        return redirect("grade")

    return render(request, "grade.html", {
        "grades": grades
    })

@login_required(login_url="/")
def editgrade(request, id):
    grade = get_object_or_404(Grade, id=id)

    if request.method == "POST":
        name = request.POST.get("grade_name")

        if Grade.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, "Grade already exists!")
        else:
            grade.name = name
            grade.save()
            messages.success(request, "Grade Updated Successfully!")

        return redirect("grade")
    return redirect("grade")

@login_required(login_url="/")
def deletegrade(request, id):
    grade = get_object_or_404(Grade, id=id)
    grade.delete()

    messages.success(request, "Grade Deleted Successfully!")
    return redirect("grade")

@login_required(login_url="/")
def country(request):
    countries = Country.objects.all().order_by("-id")

    if request.method == "POST":
        country_name = request.POST.get("country_name")

        if Country.objects.filter(name__iexact=country_name).exists():
            messages.error(request, "Country already exists!")
            return redirect("country")
        
        Country.objects.create(name=country_name)

        messages.success(request, "Country Added Successfully!")
        return redirect("country")

    return render(request, "country.html", {
        "countries": countries
    })

@login_required(login_url="/")
def editcountry(request, id):
    country = get_object_or_404(Country, id=id)

    if request.method == "POST":
        name = request.POST.get("country_name")

        if Country.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, "Country already exists!")
        else:
            country.name = name
            country.save()
            messages.success(request, "Country Updated Successfully!")

        return redirect("country")
    return redirect("country")

@login_required(login_url="/")
def deletecountry(request, id):
    country = get_object_or_404(Country, id=id)
    country.delete()

    messages.success(request, "Country Deleted Successfully!")
    return redirect("country")

@login_required(login_url="/")
def state(request):
    states = State.objects.select_related('country').all().order_by("-id")
    countries = Country.objects.all()

    if request.method == "POST":
        country_id = request.POST.get("country")
        state_name = request.POST.get("statename")

        # Validation
        if not country_id:
            messages.error(request, "Please select a country.")
            return redirect("state")

        # Duplicate check inside same country
        if State.objects.filter(
            country_id=country_id,
            name__iexact=state_name
        ).exists():
            messages.error(request, "State already exists in this country!")
            return redirect("state")

        State.objects.create(
            country_id=country_id,
            name=state_name
        )

        messages.success(request, "State Added Successfully!")
        return redirect("state")

    return render(request, "states.html", {
        "states": states,
        "countries": countries
    })

@login_required(login_url="/")
def editstate(request, id):
    state = get_object_or_404(State, id=id)

    if request.method == "POST":
        country_id = request.POST.get("country")
        state_name = request.POST.get("statename")

        # Validation
        if not country_id:
            messages.error(request, "Please select a country.")
            return redirect("state")

        # Duplicate check
        if State.objects.filter(
            country_id=country_id,
            name__iexact=state_name
        ).exclude(id=id).exists():
            messages.error(request, "State already exists in this country!")
            return redirect("state")

        state.country_id = country_id
        state.name = state_name
        state.save()

        messages.success(request, "State Updated Successfully!")
        return redirect("state")

    return redirect("state")

@login_required(login_url="/")
def deletestate(request, id):
    state = get_object_or_404(State, id=id)
    state.delete()

    messages.success(request, "State Deleted Successfully!")
    return redirect("state")


@login_required(login_url="/")
def city(request):
    cities = City.objects.select_related('state', 'state__country').all().order_by("-id")
    states = State.objects.select_related('country').all()

    if request.method == "POST":
        state_id = request.POST.get("state")
        city_name = request.POST.get("cityname")

        if not state_id:
            messages.error(request, "Please select a state.")
            return redirect("city")

        # Duplicate check inside same state
        if City.objects.filter(
            state_id=state_id,
            name__iexact=city_name
        ).exists():
            messages.error(request, "City already exists in this state!")
            return redirect("city")

        City.objects.create(
            state_id=state_id,
            name=city_name
        )

        messages.success(request, "City Added Successfully!")
        return redirect("city")

    return render(request, "city.html", {
        "cities": cities,
        "states": states
    })



@login_required(login_url="/")
def editcity(request, id):
    city = get_object_or_404(City, id=id)

    if request.method == "POST":
        state_id = request.POST.get("state")
        city_name = request.POST.get("cityname")

        if not state_id:
            messages.error(request, "Please select a state.")
            return redirect("city")

        # Duplicate check inside same state
        if City.objects.filter(
            state_id=state_id,
            name__iexact=city_name
        ).exclude(id=id).exists():
            messages.error(request, "City already exists in this state!")
            return redirect("city")

        city.state_id = state_id
        city.name = city_name
        city.save()

        messages.success(request, "City Updated Successfully!")
        return redirect("city")

    return redirect("city")

@login_required(login_url="/")
def deletecity(request, id):
    city = get_object_or_404(City, id=id)
    city.delete()

    messages.success(request, "City Deleted Successfully!")
    return redirect("city")

@login_required(login_url="/")
def area(request):
    areas = Area.objects.select_related(
        'city',
        'city__state',
        'city__state__country'
    ).all().order_by("-id")

    cities = City.objects.select_related(
        'state',
        'state__country'
    ).all()

    if request.method == "POST":
        city_id = request.POST.get("city")
        area_name = request.POST.get("areaname")

        if not city_id:
            messages.error(request, "Please select a city.")
            return redirect("area")

        # Duplicate check inside same city
        if Area.objects.filter(
            city_id=city_id,
            name__iexact=area_name
        ).exists():
            messages.error(request, "Area already exists in this city!")
            return redirect("area")

        Area.objects.create(
            city_id=city_id,
            name=area_name
        )

        messages.success(request, "Area Added Successfully!")
        return redirect("area")

    return render(request, "area.html", {
        "areas": areas,
        "cities": cities
    })

@login_required(login_url="/")
def editarea(request, id):
    area = get_object_or_404(Area, id=id)

    if request.method == "POST":
        city_id = request.POST.get("city")
        area_name = request.POST.get("areaname")

        if not city_id:
            messages.error(request, "Please select a city.")
            return redirect("area")

        # Duplicate check inside same city
        if Area.objects.filter(
            city_id=city_id,
            name__iexact=area_name
        ).exclude(id=id).exists():
            messages.error(request, "Area already exists in this city!")
            return redirect("area")

        area.city_id = city_id
        area.name = area_name
        area.save()

        messages.success(request, "Area Updated Successfully!")
        return redirect("area")

    return redirect("area")

@login_required(login_url="/")
def deletearea(request, id):
    area = get_object_or_404(Area, id=id)
    area.delete()

    messages.success(request, "Area Deleted Successfully!")
    return redirect("area")

@login_required(login_url="/")
def pincode(request):

    pincodes = Pincode.objects.select_related(
        'area',
        'area__city',
        'area__city__state',
        'area__city__state__country'
    ).all().order_by("-id")

    areas = Area.objects.select_related(
        'city',
        'city__state',
        'city__state__country'
    ).all()

    if request.method == "POST":
        area_id = request.POST.get("area")
        pincode_value = request.POST.get("pincode")

        if not area_id:
            messages.error(request, "Please select an area.")
            return redirect("pincode")

        # Duplicate check inside same area
        if Pincode.objects.filter(
            area_id=area_id,
            code__iexact=pincode_value
        ).exists():
            messages.error(request, "Pincode already exists in this area!")
            return redirect("pincode")

        Pincode.objects.create(
            area_id=area_id,
            code=pincode_value
        )

        messages.success(request, "Pincode Added Successfully!")
        return redirect("pincode")

    return render(request, "pincode.html", {
        "pincodes": pincodes,
        "areas": areas
    })

@login_required(login_url="/")
def editpincode(request, id):
    pincode = get_object_or_404(Pincode, id=id)

    if request.method == "POST":
        area_id = request.POST.get("area")
        pincode_value = request.POST.get("pincode")

        if not area_id:
            messages.error(request, "Please select an area.")
            return redirect("pincode")

        # Duplicate check inside same area
        if Pincode.objects.filter(
            area_id=area_id,
            code__iexact=pincode_value
        ).exclude(id=id).exists():
            messages.error(request, "Pincode already exists in this area!")
            return redirect("pincode")

        pincode.area_id = area_id
        pincode.code = pincode_value
        pincode.save()

        messages.success(request, "Pincode Updated Successfully!")
        return redirect("pincode")

    return redirect("pincode")

@login_required(login_url="/")
def deletepincode(request, id):
    pincode = get_object_or_404(Pincode, id=id)
    pincode.delete()

    messages.success(request, "Pincode Deleted Successfully!")
    return redirect("pincode")

@login_required(login_url="/")
def relationship(request):

    relationships = Relationship.objects.all().order_by("-id")

    # ADD
    if request.method == "POST" and "add_relationship" in request.POST:
        name = request.POST.get("relationshipname")

        if Relationship.objects.filter(name__iexact=name).exists():
            messages.error(request, "Relationship already exists!")
        else:
            Relationship.objects.create(name=name)
            messages.success(request, "Relationship Added Successfully!")

        return redirect("relationship")

    return render(request, "relationship.html", {
        "relationships": relationships
    })



@login_required(login_url="/")# EDIT
def editrelationship(request, id):
    relation = get_object_or_404(Relationship, id=id)

    if request.method == "POST":
        name = request.POST.get("relationshipname")

        if Relationship.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, "Relationship already exists!")
        else:
            relation.name = name
            relation.save()
            messages.success(request, "Relationship Updated Successfully!")

        return redirect("relationship")

    return redirect("relationship")



@login_required(login_url="/")# DELETE
def deleterelationship(request, id):
    relation = get_object_or_404(Relationship, id=id)
    relation.delete()
    messages.success(request, "Relationship Deleted Successfully!")
    return redirect("relationship")


@login_required(login_url="/")
def chronic(request):

    diseases = ChronicDisease.objects.all().order_by("-id")

    # ADD
    if request.method == "POST" and "add_disease" in request.POST:
        name = request.POST.get("diseasename")

        if ChronicDisease.objects.filter(name__iexact=name).exists():
            messages.error(request, "Chronic Disease already exists!")
        else:
            ChronicDisease.objects.create(name=name)
            messages.success(request, "Chronic Disease Added Successfully!")

        return redirect("chronic")

    return render(request, "chronic.html", {
        "diseases": diseases
    })



@login_required(login_url="/")# EDIT
def editchronic(request, id):
    disease = get_object_or_404(ChronicDisease, id=id)

    if request.method == "POST":
        name = request.POST.get("diseasename")

        if ChronicDisease.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, "Chronic Disease already exists!")
        else:
            disease.name = name
            disease.save()
            messages.success(request, "Chronic Disease Updated Successfully!")

        return redirect("chronic")

    return redirect("chronic")



@login_required(login_url="/")# DELETE
def deletechronic(request, id):
    disease = get_object_or_404(ChronicDisease, id=id)
    disease.delete()
    messages.success(request, "Chronic Disease Deleted Successfully!")
    return redirect("chronic")


@login_required(login_url="/")
def qualification(request):

    qualifications = Qualification.objects.all().order_by("-id")

    # ADD
    if request.method == "POST" and "add_qualification" in request.POST:
        name = request.POST.get("qualificationname")

        if Qualification.objects.filter(name__iexact=name).exists():
            messages.error(request, "Qualification already exists!")
        else:
            Qualification.objects.create(name=name)
            messages.success(request, "Qualification Added Successfully!")

        return redirect("qualification")

    return render(request, "qualification.html", {
        "qualifications": qualifications
    })



@login_required(login_url="/")# EDIT
def editqualification(request, id):
    qualification = get_object_or_404(Qualification, id=id)

    if request.method == "POST":
        name = request.POST.get("qualificationname")

        if Qualification.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, "Qualification already exists!")
        else:
            qualification.name = name
            qualification.save()
            messages.success(request, "Qualification Updated Successfully!")

        return redirect("qualification")

    return redirect("qualification")



@login_required(login_url="/")# DELETE
def deletequalification(request, id):
    qualification = get_object_or_404(Qualification, id=id)
    qualification.delete()
    messages.success(request, "Qualification Deleted Successfully!")
    return redirect("qualification")

@login_required(login_url="/")
def occupation(request):

    occupations = Occupation.objects.all().order_by("-id")

    # ADD
    if request.method == "POST" and "add_occupation" in request.POST:
        name = request.POST.get("occupationname")

        if Occupation.objects.filter(name__iexact=name).exists():
            messages.error(request, "Occupation already exists!")
        else:
            Occupation.objects.create(name=name)
            messages.success(request, "Occupation Added Successfully!")

        return redirect("occupation")

    return render(request, "occupation.html", {
        "occupations": occupations
    })



@login_required(login_url="/")# EDIT
def editoccupation(request, id):
    occupation = get_object_or_404(Occupation, id=id)

    if request.method == "POST":
        name = request.POST.get("occupationname")

        if Occupation.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, "Occupation already exists!")
        else:
            occupation.name = name
            occupation.save()
            messages.success(request, "Occupation Updated Successfully!")

        return redirect("occupation")

    return redirect("occupation")



@login_required(login_url="/")# DELETE
def deleteoccupation(request, id):
    occupation = get_object_or_404(Occupation, id=id)
    occupation.delete()
    messages.success(request, "Occupation Deleted Successfully!")
    return redirect("occupation")


@login_required(login_url="/")
def income(request):

    incomes = Income.objects.all().order_by("-id")

    # ADD
    if request.method == "POST" and "add_income" in request.POST:
        name = request.POST.get("incomename")

        if Income.objects.filter(name__iexact=name).exists():
            messages.error(request, "Income already exists!")
        else:
            Income.objects.create(name=name)
            messages.success(request, "Income Added Successfully!")

        return redirect("income")

    return render(request, "income.html", {
        "incomes": incomes
    })



@login_required(login_url="/")# EDIT
def editincome(request, id):
    income = get_object_or_404(Income, id=id)

    if request.method == "POST":
        name = request.POST.get("incomename")

        if Income.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, "Income already exists!")
        else:
            income.name = name
            income.save()
            messages.success(request, "Income Updated Successfully!")

        return redirect("income")

    return redirect("income")



@login_required(login_url="/")# DELETE
def deleteincome(request, id):
    income = get_object_or_404(Income, id=id)
    income.delete()
    messages.success(request, "Income Deleted Successfully!")
    return redirect("income")


@login_required(login_url="/")
def bloodgroup(request):

    bloodgroups = BloodGroup.objects.all().order_by("-id")

    # ADD
    if request.method == "POST" and "add_bloodgroup" in request.POST:
        name = request.POST.get("bloodgroupname")

        if BloodGroup.objects.filter(name__iexact=name).exists():
            messages.error(request, "Blood Group already exists!")
        else:
            BloodGroup.objects.create(name=name)
            messages.success(request, "Blood Group Added Successfully!")

        return redirect("bloodgroup")

    return render(request, "bloodgroup.html", {
        "bloodgroups": bloodgroups
    })



@login_required(login_url="/")# EDIT
def editbloodgroup(request, id):
    bloodgroup = get_object_or_404(BloodGroup, id=id)

    if request.method == "POST":
        name = request.POST.get("bloodgroupname")

        if BloodGroup.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, "Blood Group already exists!")
        else:
            bloodgroup.name = name
            bloodgroup.save()
            messages.success(request, "Blood Group Updated Successfully!")

        return redirect("bloodgroup")

    return redirect("bloodgroup")



@login_required(login_url="/")# DELETE
def deletebloodgroup(request, id):
    bloodgroup = get_object_or_404(BloodGroup, id=id)
    bloodgroup.delete()
    messages.success(request, "Blood Group Deleted Successfully!")
    return redirect("bloodgroup")

@login_required(login_url="/")
def language(request):

    languages = Language.objects.all().order_by("-id")

    # ADD
    if request.method == "POST" and "add_language" in request.POST:
        name = request.POST.get("languagename")

        if Language.objects.filter(name__iexact=name).exists():
            messages.error(request, "Language already exists!")
        else:
            Language.objects.create(name=name)
            messages.success(request, "Language Added Successfully!")

        return redirect("language")

    return render(request, "language.html", {
        "languages": languages
    })



@login_required(login_url="/")# EDIT
def editlanguage(request, id):
    language = get_object_or_404(Language, id=id)

    if request.method == "POST":
        name = request.POST.get("languagename")

        if Language.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, "Language already exists!")
        else:
            language.name = name
            language.save()
            messages.success(request, "Language Updated Successfully!")

        return redirect("language")

    return redirect("language")



@login_required(login_url="/")# DELETE
def deletelanguage(request, id):
    language = get_object_or_404(Language, id=id)
    language.delete()
    messages.success(request, "Language Deleted Successfully!")
    return redirect("language")





@login_required(login_url="/")    
def department(request):
    return render(request, 'department.html')



@login_required(login_url="/")
def health(request):

    healths = OverallHealth.objects.all().order_by("-id")

    # ADD
    if request.method == "POST" and "add_health" in request.POST:
        name = request.POST.get("healthname")

        if OverallHealth.objects.filter(name__iexact=name).exists():
            messages.error(request, "Health status already exists!")
        else:
            OverallHealth.objects.create(name=name)
            messages.success(request, "Health status added successfully!")

        return redirect("health")

    return render(request, "health.html", {
        "healths": healths
    })



@login_required(login_url="/")# EDIT
def edithealth(request, id):
    health = get_object_or_404(OverallHealth, id=id)

    if request.method == "POST":
        name = request.POST.get("healthname")

        if OverallHealth.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, "Health status already exists!")
        else:
            health.name = name
            health.save()
            messages.success(request, "Health status updated successfully!")

        return redirect("health")

    return redirect("health")



@login_required(login_url="/")# DELETE
def deletehealth(request, id):
    health = get_object_or_404(OverallHealth, id=id)
    health.delete()
    messages.success(request, "Health status deleted successfully!")
    return redirect("health")

@login_required(login_url="/")
def sports(request):

    sports = Sports.objects.all().order_by("-id")

    # ADD
    if request.method == "POST" and "add_sports" in request.POST:
        name = request.POST.get("sportsname")

        if Sports.objects.filter(name__iexact=name).exists():
            messages.error(request, "Sports already exists!")
        else:
            Sports.objects.create(name=name)
            messages.success(request, "Sports added successfully!")

        return redirect("sports")

    return render(request, "sports.html", {
        "sports": sports
    })



@login_required(login_url="/")# EDIT
def editsports(request, id):
    sports = get_object_or_404(Sports, id=id)

    if request.method == "POST":
        name = request.POST.get("sportsname")

        if Sports.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, "Sports already exists!")
        else:
            sports.name = name
            sports.save()
            messages.success(request, "Sports updated successfully!")

        return redirect("sports")

    return redirect("sports")



@login_required(login_url="/")# DELETE
def deletesports(request, id):
    sports = get_object_or_404(Sports, id=id)
    sports.delete()
    messages.success(request, "Sports deleted successfully!")
    return redirect("sports")


@login_required(login_url="/")
def sector(request):

    sector = Sectors.objects.all().order_by("-id")

    # ADD
    if request.method == "POST" and "add_sector" in request.POST:
        name = request.POST.get("sectorname")

        if Sectors.objects.filter(name__iexact=name).exists():
            messages.error(request, "Sector already exists!")
        else:
            Sectors.objects.create(name=name)
            messages.success(request, "Sector added successfully!")

        return redirect("sector")

    return render(request, "sector.html", {
        "sector": sector
    })



@login_required(login_url="/")# EDIT
def editsector(request, id):
    sector = get_object_or_404(Sectors, id=id)

    if request.method == "POST":
        name = request.POST.get("sectorname")

        if Sectors.objects.filter(name__iexact=name).exclude(id=id).exists():
            messages.error(request, "Sector already exists!")
        else:
            sector.name = name
            sector.save()
            messages.success(request, "Sector updated successfully!")

        return redirect("sector")

    return redirect("sector")



@login_required(login_url="/")# DELETE
def deletsector(request, id):
    sector = get_object_or_404(Sectors, id=id)
    sector.delete()
    messages.success(request, "Sector deleted successfully!")
    return redirect("sector")



@login_required(login_url="/")
def sitesettings(request):

    site = SiteSetting.objects.first()

    if request.method == "POST":

        site_name = request.POST.get("site_name")
        address = request.POST.get("address")
        contact_no = request.POST.get("contact_no")
        contact_email = request.POST.get("contact_email")
        logo = request.FILES.get("logo")

        if site:
            site.site_name = site_name
            site.address = address
            site.contact_no = contact_no
            site.contact_email = contact_email

            if logo:
                site.logo = logo

            site.save()
            messages.success(request, "Site Settings Updated Successfully!")

        else:
            SiteSetting.objects.create(
                site_name=site_name,
                address=address,
                contact_no=contact_no,
                contact_email=contact_email,
                logo=logo
            )

            messages.success(request, "Site Settings Saved Successfully!")

        return redirect("sitesettings")

    return render(request, "site.html", {
        "site": site
    })

@login_required(login_url="/")
def news(request):

    news = News.objects.first()   # only one news record

    if request.method == "POST":
        content = request.POST.get("content")

        if news:
            news.content = content
            news.save()
            messages.success(request, "News Updated Successfully!")
        else:
            News.objects.create(content=content)
            messages.success(request, "News Added Successfully!")

        return redirect("news")

    return render(request, "news.html", {
        "news": news
    })


@login_required(login_url="/")
def allocatevolunteer(request):

    volunteers = AddUser.objects.filter(role="volunteer")

    if request.method == "POST":
        print(request.POST)

        volunteer_id = request.POST.get("volunteer")

        volunteer = AddUser.objects.get(id=int(volunteer_id))

        VolunteerAllocation.objects.create(
            volunteer=volunteer,
            start_date=request.POST.get("start_date"), # Updated
            end_date=request.POST.get("end_date"),
            start_time=request.POST.get("start_time"),
            end_time=request.POST.get("end_time"),
            allocated_by=request.user
        )

        messages.success(request, "Volunteer Allocated Successfully")
        return redirect("allocatevolunteer")

    return render(request, "allocate-volunteer.html", {
        "volunteers": volunteers
    })


@login_required(login_url="/")
def allocatelist(request):

    allocations = VolunteerAllocation.objects.select_related(
        'volunteer',
        'allocated_by'
    ).order_by("-id")

    data = []

    now = timezone.now().time()

    for a in allocations:

        if a.end_time and a.end_time < now:
            status = "Expired"
        else:
            status = "Active"

        data.append({
            "id": a.id,
            "user_id": a.volunteer.employee_id,
            "name": a.volunteer.username,
            "start_date": a.start_date,
            "end_date": a.end_date,
            "start": a.start_time,
            "end": a.end_time,
            "assigned": a.allocated_by.username if a.allocated_by else "-",
            "status": status
        })

    return render(request, "allocate-list.html", {
        "allocations": data })

@login_required(login_url="/")   
def get_allocation_data(request, id):
    try:
        a = VolunteerAllocation.objects.get(id=id)
        data = {
            "id": a.id,
            "volunteer_id": a.volunteer.id,
            "volunteer_code": a.volunteer.employee_id,
            "volunteer_name": a.volunteer.username,
            "start_date": a.start_date.strftime("%Y-%m-%d"),
            "end_date": a.end_date.strftime("%Y-%m-%d"),
            "start_time": a.start_time.strftime("%H:%M") if a.start_time else "",
            "end_time": a.end_time.strftime("%H:%M") if a.end_time else "",
        }
        return JsonResponse(data)
    except VolunteerAllocation.DoesNotExist:
        return JsonResponse({"error": "Allocation not found"}, status=404)


@csrf_exempt
@login_required(login_url="/")
def update_allocation(request):
    if request.method == "POST":
        id = request.POST.get("id")
        a = get_object_or_404(VolunteerAllocation, id=id)

        a.start_date = request.POST.get("start_date")
        a.end_date = request.POST.get("end_date")
        a.start_time = request.POST.get("start_time") or None
        a.end_time = request.POST.get("end_time") or None
        
        a.save()
        return JsonResponse({"success": True})
    return JsonResponse({"success": False})



@login_required(login_url="/")
def activeallocation(request):
    return render(request, 'active-allocation.html')

@login_required(login_url="/")
def expiredallocation(request):
    return render(request, 'expired-allocation.html')

from django.http import JsonResponse

@login_required(login_url="/")
def get_location_by_pincode(request):

    pincode_id = request.GET.get("pincode_id")

    try:
        pincode = Pincode.objects.select_related(
            "area",
            "area__city",
            "area__city__state"
        ).get(id=pincode_id)

        data = {
            "area_id": pincode.area.id,
            "area_name": pincode.area.name,
            "city": pincode.area.city.name,
            "state": pincode.area.city.state.name,
        }

        return JsonResponse(data)

    except Pincode.DoesNotExist:
        return JsonResponse({"error": "Invalid Pincode"})

@login_required(login_url="/")
def addfamily(request):

    # Generate Next Family ID
    last_family = Family.objects.order_by('-id').first()

    if last_family:
        num = int(last_family.family_id[3:]) + 1
    else:
        num = 1

    next_family_id = f"HTF{num:04d}"

    areas = Area.objects.all()
    pincodes = Pincode.objects.all()
    citys = City.objects.all()
    states = State.objects.all()


    if request.method == "POST":

        # Resolve Area from Name provided in POST (since it's now a text input)
        area_name = (request.POST.get("area") or "").title()
        pincode_id = request.POST.get("pincode")
        pincode_obj = Pincode.objects.select_related("area__city").get(id=pincode_id)
        city_obj = pincode_obj.area.city
        
        # Get or create Area within that city
        area_obj, created = Area.objects.get_or_create(
            city=city_obj, 
            name=area_name
        )

        # CREATE FAMILY
        family = Family.objects.create(
            family_id=request.POST.get("family_id"),
            head_name=(request.POST.get("family_head_name") or "").title(),
            aadhar_no=request.POST.get("aadhar_no"),
            mobile_no=request.POST.get("mobile_no"),
            door_no=(request.POST.get("door_no") or "").title(),
            apartment_name=(request.POST.get("apartment_name") or "").title(),
            # flat_no=request.POST.get("flat_no"),
            # floor_no=request.POST.get("floor_no"),
            street_name=(request.POST.get("street_name") or "").title(),
            # road_name=request.POST.get("road_name"),
            area=area_obj,
            pincode=pincode_obj.code, # or pincode_id? Let's check model.
            landline=request.POST.get("landline"),
            residential_status=request.POST.get("residential_status"),
            ration_card=request.FILES.get("ration_card"),
            no_of_members=request.POST.get("member_count"),
            created_by=request.user
        )
    
        # -----------------------------
        # 1️⃣ ADD FAMILY HEAD AS MEMBER
        # -----------------------------

        Member.objects.create(
            family=family,
            name=(request.POST.get("family_head_name") or "").title(),
            member_type="Adult"   # default (you can change if needed)
        )

        # -----------------------------
        # 2️⃣ ADD OTHER MEMBERS
        # -----------------------------
        member_count = int(request.POST.get("member_count") or 0)

        for i in range(1, member_count + 1):

            name = (request.POST.get(f"member_name_{i}") or "").title()
            mtype = request.POST.get(f"member_type_{i}")

            if name:


                Member.objects.create(
                    family=family,
                    name=name,
                    member_type=mtype
                )

        messages.success(request, "Family Added Successfully")
        return redirect("/addfamily/")

    return render(request, "add-family.html", {
        "areas": areas,
        "pincodes": pincodes,
        "next_family_id": next_family_id
    })

# @login_required(login_url="/")
# def addfamily(request):

#     # Generate Next Family ID
#     last_family = Family.objects.order_by('-id').first()

#     if last_family:
#         num = int(last_family.family_id[3:]) + 1
#     else:
#         num = 1

#     next_family_id = f"HTF{num:04d}"

#     areas = Area.objects.all()

#     if request.method == "POST":

#         family = Family.objects.create(
#             family_id=request.POST.get("family_id"),
#             head_name=request.POST.get("family_head_name"),
#             aadhar_no=request.POST.get("aadhar_no"),
#             mobile_no=request.POST.get("mobile_no"),
#             door_no=request.POST.get("door_no"),
#             apartment_name=request.POST.get("apartment_name"),
#             flat_no=request.POST.get("flat_no"),
#             floor_no=request.POST.get("floor_no"),
#             street_name=request.POST.get("street_name"),
#             road_name=request.POST.get("road_name"),
#             area_id=request.POST.get("area"),
#             pincode=request.POST.get("pincode"),
#             landline=request.POST.get("landline"),
#             residential_status=request.POST.get("residential_status"),
#             ration_card=request.FILES.get("ration_card"),
#             no_of_members=request.POST.get("member_count"),
#             created_by=request.user
#         )

#         member_count = int(request.POST.get("member_count"))

#         for i in range(1, member_count + 1):

#             name = request.POST.get(f"member_name_{i}")
#             mtype = request.POST.get(f"member_type_{i}")

#             if name:

#                 last_member = Member.objects.order_by('-id').first()

#                 if last_member:
#                     num = int(last_member.member_id[3:]) + 1
#                 else:
#                     num = 1

#                 member_id = f"HTM{num:04d}"

#                 Member.objects.create(
#                     member_id=member_id,
                # Member.objects.create(
#                     family=family,
#                     name=name,
#                     member_type=mtype
#                 )

#         messages.success(request, "Family Added Successfully")
#         return redirect("/addfamily/")

#     return render(request, "add-family.html", {
#         "areas": areas,
#         "next_family_id": next_family_id
#     })

@login_required(login_url="/")
def familylist(request):

    families = Family.objects.all().order_by("-id")

    return render(request, "family-list.html", {
        "families": families
    })

@login_required(login_url="/")
def editfamily(request,id):

    family = get_object_or_404(Family,id=id)

    if request.method == "POST":

        # Apply Title Case to all text fields
        family.head_name = (request.POST.get("family_head_name") or "").title()
        family.aadhar_no = request.POST.get("aadhar_no")
        family.mobile_no = request.POST.get("mobile_no")
        family.door_no = (request.POST.get("door_no") or "").title()
        family.apartment_name = (request.POST.get("apartment_name") or "").title()
        family.street_name = (request.POST.get("street_name") or "").title()
        
        # Pincode and Area logic
        pincode_val = request.POST.get("pincode")
        area_name = (request.POST.get("area") or "").title()
        family.pincode = pincode_val
        family.landline = request.POST.get("landline")
        family.residential_status = request.POST.get("residential_status")

        # Attempt to resolve the Area object
        try:
            pincode_obj = Pincode.objects.filter(code=pincode_val).first()
            if pincode_obj:
                city_obj = pincode_obj.area.city
                # Get or create an Area within this city using the title-cased name
                area_obj, created = Area.objects.get_or_create(city=city_obj, name=area_name)
                family.area = area_obj
            else:
                # If pincode not found, use the current city of the family member's area
                area_obj, created = Area.objects.get_or_create(city=family.area.city, name=area_name)
                family.area = area_obj
        except Exception:
            pass

        if request.FILES.get("ration_card"):
            family.ration_card = request.FILES.get("ration_card")

        family.save()
        messages.success(request,"Family updated successfully")

    return redirect("familylist")

@login_required(login_url="/")
def deletefamily(request, id):
    family = get_object_or_404(Family, id=id)
    # Delete all members in this family first
    family.members.all().delete()
    family.delete()
    messages.success(request, "Family and all its members deleted successfully.")
    return redirect("familylist")

@login_required(login_url="/")
def addmember(request,id):

    family = Family.objects.get(id=id)

    if request.method == "POST":

        names = request.POST.getlist("member_name[]")
        types = request.POST.getlist("member_type[]")

        for name, mtype in zip(names,types):

            if name:
                name = name.title()


                Member.objects.create(
                    family=family,
                    name=name,
                    member_type=mtype
                )

        return redirect("familylist")




@login_required(login_url="/")
def update_member_details(request, id):
    member = get_object_or_404(Member, id=id)
    print(member)

    if request.method == "POST":
        print(request.POST)  #


        def get_val(field_name, is_fk=False):
            val = request.POST.get(field_name)
            if not val or val == "Not Willing":
                return None
            return val

        # -----------------------------
        # ADULT BLOCK
        # -----------------------------
        if member.member_type == "Adult":
            # Basic Info
            member.name = request.POST.get("name")
            member.aadhaar_number = request.POST.get("aadhaar_number")
            member.father_name = request.POST.get("father_name")
            member.relationship_id = request.POST.get("relationship") or None
            member.alias = request.POST.get("alias")
            member.gender = request.POST.get("gender")
            dob = request.POST.get("date_of_birth")
            member.date_of_birth = dob if dob else None

            # Contact
            member.email_id = request.POST.get("email_id")
            member.mobile = request.POST.get("mobile")
            member.whatsapp = request.POST.get("whatsapp")
            
            member.marital_status = request.POST.get("marital_status")
            member.qualification_id = request.POST.get("academic_qualification") or None
            member.diploma_degree = request.POST.get("diploma_degree")
            member.languages_speak =request.POST.get("languages_speak")
            member.languages_read =request.POST.get("languages_read")
            member.languages_write =request.POST.get("languages_write")
            member.computer_proficiency = request.POST.get("computer_proficiency")
            member.occupation_id = request.POST.get("occupation") or None
            member.company_name = request.POST.get("company_name")
            member.sector = request.POST.get("sector")
            member.industry = request.POST.get("industry")
            member.designation = request.POST.get("designation")
            member.department = request.POST.get("department")
            member.income_id = request.POST.get("annual_income") or None
            member.skills = request.POST.get("skills")
            member.blood_group_id = request.POST.get("blood_group") or None
            member.overall_health_id = request.POST.get("overall_health") or None
            member.chronic_disease_id = request.POST.get("chronic_disease") or None
            member.disability = request.POST.get("disability")
            member.activity_level = request.POST.get("activity_level")
            member.govt_health_insurance = request.POST.get("govt_ins") == "Yes"
            member.private_health_insurance = request.POST.get("private_ins") == "Yes" 
            member.husainy_trust_member = request.POST.get("husainy_trust_member") == "on"
            member.membership_number = request.POST.get("membership_number")
            member.interested_in_digital_directory = request.POST.get("interested_in_digital_directory") == "on"
            member.listing_in_community_digital_pages = request.POST.get("listing_in_community_digital_pages") == "on"
            member.product_service_listing_in_yellow_pages = request.POST.get("yellowpages")
            if request.FILES.get("photo"):
                member.photo = request.FILES.get("photo")
            
            
        # -----------------------------
        # STUDENT BLOCK
        # -----------------------------
        elif member.member_type == "Student":
            # Basic Info
            
            member.name = request.POST.get("name")
            member.aadhaar_number = request.POST.get("aadhaar_number")
            member.relationship_id = request.POST.get("relationship") or None
            member.alias = request.POST.get("alias")
            member.father_name = request.POST.get("father_name")
            member.gender = request.POST.get("gender")
            dob = request.POST.get("date_of_birth")
            member.date_of_birth = dob if dob else None

            member.current_education_status = request.POST.get("current_education_status")
            member.school_college_institute_name = request.POST.get("school_college_institute_name")
            member.grade_year = request.POST.get("grade_year")
            member.annual_academic_fees = request.POST.get("annual_academic_fees")
            member.fees_payment = request.POST.get("fees_payment")
            member.languages_speak =request.POST.get("languages_speak")
            member.languages_read =request.POST.get("languages_read")
            member.languages_write =request.POST.get("languages_write")
            member.computer_proficiency = request.POST.get("computer_proficiency")
            member.sports = request.POST.get("sports")
            member.hobbies = request.POST.get("hobbies")
            member.career_goal = request.POST.get("career_goal")
            member.holy_koran_reading = request.POST.get("holy_koran_reading")
            member.deeniyath = request.POST.get("deeniyath")
            member.blood_group_id = request.POST.get("blood_group") or None
            member.overall_health_id = request.POST.get("overall_health") or None
            member.chronic_disease_id = request.POST.get("chronic_disease") or None
            member.disability = request.POST.get("disability")
            member.govt_health_insurance = request.POST.get("govt_ins") == "Yes"
            member.private_health_insurance = request.POST.get("private_ins") == "Yes"

            if request.FILES.get("photo"):
                member.photo = request.FILES.get("photo")

        # -----------------------------
        # BABY BLOCK
        # -----------------------------
        elif member.member_type == "Baby":
            # All fields available for Baby
            member.name = request.POST.get("name")
            member.aadhaar_number = request.POST.get("aadhaar_number")
            member.relationship_id = request.POST.get("relationship") or None
            member.alias = request.POST.get("alias")
            member.father_name = request.POST.get("father_name")
            member.gender = request.POST.get("gender")
            
            dob = request.POST.get("date_of_birth")
            member.date_of_birth = dob if dob else None

            member.blood_group_id = request.POST.get("blood_group") or None
            member.overall_health_id = request.POST.get("overall_health") or None
            member.govt_health_insurance = request.POST.get("govt_ins") == "Yes"
            member.private_health_insurance = request.POST.get("private_ins") == "Yes"
            
            
            if request.FILES.get("photo"):
                member.photo = request.FILES.get("photo")

        # -----------------------------
        # SAVE CHANGES
        # -----------------------------
        if request.POST.get("is_final") == "True":
            member.is_updated = True
        
        member.save()

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({"status": "success", "is_updated": member.is_updated, "member_id": member.id})

        return redirect("/memberlist/")

    return render(request,"member-list.html",{"member":member})

@login_required(login_url="/")
def edit_member(request, id):  
    member = get_object_or_404(Member, id=id)
    relationships = Relationship.objects.all()
    bloodgroups = BloodGroup.objects.all()
    diseases = ChronicDisease.objects.all()
    qualifications = Qualification.objects.all()
    occupations = Occupation.objects.all()
    incomes = Income.objects.all()
    languages = Language.objects.all()
    healths = OverallHealth.objects.all()
    sectors = Sectors.objects.all()
    sports = Sports.objects.all()
    grades = Grade.objects.all()

    if request.method == "POST":
        print(request.POST)  #

        # -----------------------------
        # ADULT BLOCK
        # -----------------------------
        if member.member_type == "Adult":
            # Basic Info
            member.name = request.POST.get("name")
            member.aadhaar_number = request.POST.get("aadhaar_number")
            member.father_name = request.POST.get("father_name")
            member.relationship_id = request.POST.get("relationship") or None
            member.alias = request.POST.get("alias")
            member.gender = request.POST.get("gender")
            dob = request.POST.get("date_of_birth")
            member.date_of_birth = dob if dob else None

            # Contact
            member.email_id = request.POST.get("email_id")
            member.mobile = request.POST.get("mobile")
            member.whatsapp = request.POST.get("whatsapp")
            
            member.marital_status = request.POST.get("marital_status")
            member.qualification_id = request.POST.get("academic_qualification") or None
            member.diploma_degree = request.POST.get("diploma_degree")
            member.languages_speak = request.POST.get("languages_speak") or None
            member.languages_read = request.POST.get("languages_read") or None
            member.languages_write = request.POST.get("languages_write") or None
            member.computer_proficiency = request.POST.get("computer_proficiency")
            member.occupation_id = request.POST.get("occupation") or None
            member.company_name = request.POST.get("company_name")
            member.sector = request.POST.get("sector") or None
            member.industry = request.POST.get("industry")
            member.designation = request.POST.get("designation")
            member.department = request.POST.get("department")
            member.income_id = request.POST.get("annual_income") or None
            member.skills = request.POST.get("skills")
            member.blood_group_id = request.POST.get("blood_group") or None
            member.overall_health_id = request.POST.get("overall_health") or None
            member.chronic_disease_id = request.POST.get("chronic_disease") or None
            member.disability = request.POST.get("disability")
            member.activity_level = request.POST.get("activity_level")
            member.govt_health_insurance = request.POST.get("govt_health_insurance") == "True"
            member.private_health_insurance = request.POST.get("private_health_insurance") == "True" 
            member.husainy_trust_member = request.POST.get("husainy_trust_member") == "True"
            member.membership_number = request.POST.get("membership_number")
            member.interested_in_digital_directory = request.POST.get("interested_in_digital_directory") == "True"
            member.listing_in_community_digital_pages = request.POST.get("listing_in_community_digital_pages") == "True"
            member.product_service_listing_in_yellow_pages = request.POST.get("product_service_listing_in_yellow_pages")
            if request.FILES.get("photo"):
                member.photo = request.FILES.get("photo")
            
            
        # -----------------------------
        # STUDENT BLOCK
        # -----------------------------
        elif member.member_type == "Student":
            # Basic Info
            
            member.name = request.POST.get("name")
            member.aadhaar_number = request.POST.get("aadhaar_number")
            member.relationship_id = request.POST.get("relationship") or None
            member.alias = request.POST.get("alias")
            member.father_name = request.POST.get("father_name")
            member.gender = request.POST.get("gender")
            dob = request.POST.get("date_of_birth")
            member.date_of_birth = dob if dob else None

            member.current_education_status = request.POST.get("current_education_status")
            member.school_college_institute_name = request.POST.get("school_college_institute_name")
            member.grade_year = request.POST.get("grade_year") or None
            member.annual_academic_fees = request.POST.get("annual_academic_fees")
            member.fees_payment = request.POST.get("fees_payment")
            member.languages_speak = request.POST.get("languages_speak") or None
            member.languages_read = request.POST.get("languages_read") or None
            member.languages_write = request.POST.get("languages_write") or None
            member.computer_proficiency = request.POST.get("computer_proficiency")
            member.sports = request.POST.get("sports") or None
            member.hobbies = request.POST.get("hobbies")
            member.career_goal = request.POST.get("career_goal")
            member.holy_koran_reading = request.POST.get("holy_koran_reading")
            member.deeniyath = request.POST.get("deeniyath")
            member.blood_group_id = request.POST.get("blood_group") or None
            member.overall_health_id = request.POST.get("overall_health") or None
            member.chronic_disease_id = request.POST.get("chronic_disease") or None
            member.disability = request.POST.get("disability")
            member.govt_health_insurance = request.POST.get("govt_health_insurance") == "True"
            member.private_health_insurance = request.POST.get("private_health_insurance") == "True"

            if request.FILES.get("photo"):
                member.photo = request.FILES.get("photo")

        # -----------------------------
        # BABY BLOCK
        # -----------------------------
        elif member.member_type == "Baby":
            # All fields available for Baby
            member.name = request.POST.get("name")
            member.aadhaar_number = request.POST.get("aadhaar_number")
            member.relationship_id = request.POST.get("relationship") or None
            member.alias = request.POST.get("alias")
            member.father_name = request.POST.get("father_name")
            member.gender = request.POST.get("gender")
            
            dob = request.POST.get("date_of_birth")
            member.date_of_birth = dob if dob else None

            member.blood_group_id = request.POST.get("blood_group") or None
            member.overall_health_id = request.POST.get("overall_health") or None
            member.govt_health_insurance = request.POST.get("govt_health_insurance") == "True"
            member.private_health_insurance = request.POST.get("private_health_insurance") == "True"
            
            
            if request.FILES.get("photo"):
                member.photo = request.FILES.get("photo")

        # -----------------------------
        # SAVE CHANGES
        # -----------------------------
        member.is_updated = True
        member.save()
        return redirect("/completemember/")

    return render(request,"edit-member.html",{'member':member,"relationships": relationships,
        "bloodgroups": bloodgroups,
        "diseases": diseases,
        "qualifications": qualifications,
        "occupations": occupations,
        "incomes": incomes,
        "languages": languages,
        "healths": healths,
        "sectors": sectors,
        "sports": sports,
        "grades": grades,})





@login_required(login_url="/")
def memberlist(request):
    members = Member.objects.filter(is_updated=False).order_by("-id")
    relationships = Relationship.objects.all()
    bloodgroups = BloodGroup.objects.all()
    diseases = ChronicDisease.objects.all()
    qualifications = Qualification.objects.all()
    occupations = Occupation.objects.all()
    incomes = Income.objects.all()
    languages = Language.objects.all()
    healths = OverallHealth.objects.all()
    grades = Grade.objects.all()
    sports = Sports.objects.all()
    sectors = Sectors.objects.all()
    print(sports)

    return render(request, "member-list.html", {
        "members": members,
        "relationships": relationships,
        "bloodgroups": bloodgroups,
        "diseases": diseases,
        "qualifications": qualifications,
        "occupations": occupations,
        "incomes": incomes,
        "languages": languages,
        "healths": healths,
        "grades": grades,
        "sports": sports,
        "sectors": sectors,
    })


@login_required(login_url="/")
def viewmember(request,id):

    member = get_object_or_404(Member, id = id)

    return render(request,'view-member.html',{'member':member})

@login_required(login_url="/")
def completemember(request):

    members = Member.objects.filter(

        # Adult completed conditions
        Q(
            member_type="Adult",
            name__isnull=False,
            aadhaar_number__isnull=False,
            father_name__isnull=False,
            relationship__isnull=False,
            alias__isnull=False,
            gender__isnull=False,
            date_of_birth__isnull=False,
            email_id__isnull=False,
            mobile__isnull=False,
            whatsapp__isnull=False,
            marital_status__isnull=False,
            qualification__isnull=False,
            diploma_degree__isnull=False,
            languages_speak__isnull=False,
            languages_read__isnull=False,
            languages_write__isnull=False,
            computer_proficiency__isnull=False,
            occupation_id__isnull=False,
            company_name__isnull=False,
            sector__isnull=False,
            industry__isnull=False,
            designation__isnull=False,
            department__isnull=False,
            income__isnull=False,
            skills__isnull=False,
            blood_group__isnull=False,
            overall_health__isnull=False,
            chronic_disease__isnull=False,
            disability__isnull=False,
            activity_level__isnull=False,
            govt_health_insurance__isnull=False,
            private_health_insurance__isnull=False,
            husainy_trust_member__isnull=False,
            membership_number__isnull=False,
            interested_in_digital_directory__isnull=False,
            listing_in_community_digital_pages__isnull=False,
            product_service_listing_in_yellow_pages__isnull=False,
            photo__isnull=False
        )

        |

        # Student completed conditions
        Q(
            member_type="Student",
            name__isnull=False,
            aadhaar_number__isnull=False,
            relationship_id__isnull=False,
            alias__isnull=False,
            father_name__isnull=False,
            gender__isnull=False,
            date_of_birth__isnull=False,
            current_education_status__isnull=False,
            school_college_institute_name__isnull=False,
            grade_year__isnull=False,
            annual_academic_fees__isnull=False,
            fees_payment__isnull=False,
            languages_speak__isnull=False,
            languages_read__isnull=False,
            languages_write__isnull=False,
            computer_proficiency__isnull=False,
            sports__isnull=False,
            hobbies__isnull=False,
            career_goal__isnull=False,
            holy_koran_reading__isnull=False,
            deeniyath__isnull=False,
            blood_group__isnull=False,
            overall_health__isnull=False,
            chronic_disease__isnull=False,
            disability__isnull=False,
            govt_health_insurance__isnull=False,
            private_health_insurance__isnull=False,
            photo__isnull=False,
        )

        |

        # Baby completed conditions
        Q(
            member_type="Baby",
            name__isnull=False,
            relationship_id__isnull=False,
            alias__isnull=False,
            father_name__isnull=False,
            gender__isnull=False,
            date_of_birth__isnull=False,
            blood_group__isnull=False,
            overall_health__isnull=False,
            govt_health_insurance__isnull=False,
            private_health_insurance__isnull=False,
            photo__isnull=False,
        )

    )

    return render(request, "completed-member.html", {"members": members})

@login_required(login_url="/")
def incompletemember(request):

    completed_members = Member.objects.filter(

        Q(
            member_type="Adult",
            name__isnull=False,
            aadhaar_number__isnull=False,
            father_name__isnull=False,
            relationship__isnull=False,
            alias__isnull=False,
            gender__isnull=False,
            date_of_birth__isnull=False,
            email_id__isnull=False,
            mobile__isnull=False,
            whatsapp__isnull=False,
            marital_status__isnull=False,
            qualification__isnull=False,
            diploma_degree__isnull=False,
            languages_speak__isnull=False,
            languages_read__isnull=False,
            languages_write__isnull=False,
            computer_proficiency__isnull=False,
            occupation_id__isnull=False,
            company_name__isnull=False,
            sector__isnull=False,
            industry__isnull=False,
            designation__isnull=False,
            department__isnull=False,
            income__isnull=False,
            skills__isnull=False,
            blood_group__isnull=False,
            overall_health__isnull=False,
            chronic_disease__isnull=False,
            disability__isnull=False,
            activity_level__isnull=False,
            membership_number__isnull=False,
            product_service_listing_in_yellow_pages__isnull=False,
            photo__isnull=False
        )

        |

        Q(
            member_type="Student",
            name__isnull=False,
            aadhaar_number__isnull=False,
            relationship_id__isnull=False,
            alias__isnull=False,
            father_name__isnull=False,
            gender__isnull=False,
            date_of_birth__isnull=False,
            current_education_status__isnull=False,
            school_college_institute_name__isnull=False,
            grade_year__isnull=False,
            annual_academic_fees__isnull=False,
            fees_payment__isnull=False,
            languages_speak__isnull=False,
            languages_read__isnull=False,
            languages_write__isnull=False,
            computer_proficiency__isnull=False,
            sports__isnull=False,
            hobbies__isnull=False,
            career_goal__isnull=False,
            holy_koran_reading__isnull=False,
            deeniyath__isnull=False,
            blood_group__isnull=False,
            overall_health__isnull=False,
            chronic_disease__isnull=False,
            disability__isnull=False,
            photo__isnull=False,
        )

        |

        Q(
            member_type="Baby",
            name__isnull=False,
            relationship_id__isnull=False,
            alias__isnull=False,
            father_name__isnull=False,
            gender__isnull=False,
            date_of_birth__isnull=False,
            blood_group__isnull=False,
            overall_health__isnull=False,
            photo__isnull=False,
        )

    )

    incomplete_members = Member.objects.exclude(
        id__in=completed_members.values_list("id", flat=True)
    )

    return render(request, "incompleted-member.html", {"members": incomplete_members})

@login_required(login_url="/")
def completedfamily(request):
    # 1. Start with your existing field validation
    families = Family.objects.filter(
        head_name__isnull=False, aadhar_no__isnull=False,
        mobile_no__isnull=False, door_no__isnull=False,
        floor_no__isnull=False, street_name__isnull=False,
        road_name__isnull=False, area__isnull=False,
        pincode__isnull=False, residential_status__isnull=False,
        no_of_members__isnull=False
    ).exclude(
        Q(head_name="") | Q(aadhar_no="") | Q(mobile_no="") |
        Q(door_no="") | Q(floor_no="") | Q(street_name="") |
        Q(road_name="") | Q(pincode="")
    )

    # 2. Add the Member Check: 
    # EXCLUDE families where at least one member has is_updated=False
    families = families.exclude(members__is_updated=False)

    # 3. Add Area/City/State optimization for the list view
    families = families.select_related('area__city__state')

    return render(request, "completed-family.html", {"families": families})

@login_required(login_url="/")
def incompletefamily(request):
    # 1. Families with empty/null fields
    field_empty_query = (
        Q(head_name__isnull=True) | Q(head_name="") |
        Q(aadhar_no__isnull=True) | Q(aadhar_no="") |
        Q(mobile_no__isnull=True) | Q(mobile_no="") |
        Q(door_no__isnull=True) | Q(door_no="") |
        Q(floor_no__isnull=True) | Q(floor_no="") |
        Q(street_name__isnull=True) | Q(street_name="") |
        Q(road_name__isnull=True) | Q(road_name="") |
        Q(area__isnull=True) |
        Q(pincode__isnull=True) | Q(pincode="") |
        Q(residential_status__isnull=True) |
        Q(no_of_members__isnull=True)
    )

    # 2. Families where at least one member is NOT updated
    member_not_updated_query = Q(members__is_updated=False)

    
    families = Family.objects.filter(
        field_empty_query | member_not_updated_query
    ).distinct() # .distinct() is important because one family has many members

    return render(request, "incomplete-family.html", {"families": families})

@login_required(login_url="/")
def unapprovedfamily(request):
    return render(request, 'unapprove-family.html')

@login_required(login_url="/")
def useractivity(request):
    return render(request, 'user-activity.html')

@login_required(login_url="/")
def advancereport(request):
    # Universal single-field search across ALL member fields
    q = request.GET.get('q', '').strip()

    members = Member.objects.all().select_related(
        'family', 'blood_group', 'occupation', 'qualification',
        'income', 'overall_health', 'chronic_disease', 'relationship'
    )

    if q:
        members = members.filter(
            # ── Identity ───────────────────────────────────────────
            Q(name__icontains=q) |
            Q(member_id__icontains=q) |
            Q(family__family_id__icontains=q) |
            Q(alias__icontains=q) |
            Q(father_name__icontains=q) |
            Q(aadhaar_number__icontains=q) |
            Q(member_type__icontains=q) |

            # ── Personal ────────────────────────────────────────────
            Q(gender__icontains=q) |
            Q(marital_status__icontains=q) |
            Q(date_of_birth__icontains=q) |
            Q(relationship__name__icontains=q) |

            # ── Contact ─────────────────────────────────────────────
            Q(mobile__icontains=q) |
            Q(whatsapp__icontains=q) |
            Q(email_id__icontains=q) |

            # ── Health ──────────────────────────────────────────────
            Q(blood_group__name__icontains=q) |
            Q(overall_health__name__icontains=q) |
            Q(chronic_disease__name__icontains=q) |
            Q(disability__icontains=q) |
            Q(activity_level__icontains=q) |

            # ── Location ─────────────────────────────────────────────
            Q(family__area__name__icontains=q) |
            Q(family__area__city__name__icontains=q) |

            # ── Adult — Education / Work ─────────────────────────────
            Q(qualification__name__icontains=q) |
            Q(occupation__name__icontains=q) |
            Q(income__name__icontains=q) |
            Q(designation__icontains=q) |
            Q(department__icontains=q) |
            Q(company_name__icontains=q) |
            Q(skills__icontains=q) |
            Q(diploma_degree__icontains=q) |
            Q(sector__icontains=q) |
            Q(industry__icontains=q) |
            Q(membership_number__icontains=q) |
            Q(product_service_listing_in_yellow_pages__icontains=q) |

            # ── Language / Tech ──────────────────────────────────────
            Q(languages_speak__icontains=q) |
            Q(languages_read__icontains=q) |
            Q(languages_write__icontains=q) |
            Q(computer_proficiency__icontains=q) |

            # ── Student ──────────────────────────────────────────────
            Q(current_education_status__icontains=q) |
            Q(school_college_institute_name__icontains=q) |
            Q(grade_year__icontains=q) |
            Q(fees_payment__icontains=q) |
            Q(annual_academic_fees__icontains=q) |
            Q(sports__icontains=q) |
            Q(hobbies__icontains=q) |
            Q(career_goal__icontains=q) |
            Q(holy_koran_reading__icontains=q) |
            Q(deeniyath__icontains=q)
        ).distinct()

    context = {
        'members': members,
        'q': q,
        'total': members.count() if q else None,
    }
    return render(request, 'advance-report.html', context)

@login_required(login_url="/")
def usereport(request):

    selected_user = request.GET.get("user")
    users = AddUser.objects.filter(role__in=["admin", "manager"])

    reports = LoginReport.objects.filter(
        user__adduser__role__in=["admin", "manager"]
    ).select_related("user", "user__adduser").order_by("-id")

    if selected_user:
        reports = reports.filter(user_id=selected_user)

    context = {
        "users": users,
        "reports": reports,
        "selected_user": selected_user
    }

    return render(request, "user-report.html", context)

@login_required(login_url="/")
def volunteenreport(request):
    selected_user = request.GET.get("user")
    users = AddUser.objects.filter(role="volunteer")
    reports = LoginReport.objects.filter(
        user__adduser__role="volunteer"
    ).select_related("user", "user__adduser").order_by("-id")

    if selected_user:
        reports = reports.filter(user_id=selected_user)

    context = {
        "users": users,
        "reports": reports,
        "selected_user": selected_user
    }
    return render(request, 'volunteen-report.html',context)

@login_required(login_url="/")
def export_user_excel(request):

    selected_user = request.GET.get("user")

    reports = LoginReport.objects.filter(
        user__adduser__role__in=["admin", "manager"]
    ).select_related("user", "user__adduser")

    if selected_user:
        reports = reports.filter(user_id=selected_user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Report"

    ws.append([
        "Employee ID",
        "Name",
        "Date",
        "Login Time",
        "Logout Time"
    ])

    for r in reports:

        ws.append([
            r.user.adduser.employee_id,
            r.user.adduser.fullname,
            r.created_at.strftime("%d-%m-%Y"),
            r.login_time.strftime("%I:%M %p"),
            r.logout_time.strftime("%I:%M %p") if r.logout_time else ""
        ])

    response = HttpResponse(
        content_type="application/ms-excel"
    )

    response["Content-Disposition"] = "attachment; filename=user_report.xlsx"

    wb.save(response)

    return response

@login_required(login_url="/")
def export_user_pdf(request):

    selected_user = request.GET.get("user")

    reports = LoginReport.objects.filter(
        user__adduser__role__in=["admin", "manager"]
    ).select_related("user", "user__adduser")

    if selected_user:
        reports = reports.filter(user_id=selected_user)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=user_report.pdf"

    p = canvas.Canvas(response)

    y = 800

    p.drawString(50, y, "User Report")
    y -= 40

    for r in reports:

        text = f"{r.user.adduser.employee_id} - {r.user.adduser.fullname} | Login: {r.login_time} | Logout: {r.logout_time}"

        p.drawString(50, y, text)

        y -= 20

        if y < 100:
            p.showPage()
            y = 800

    p.save()

    return response

@login_required(login_url="/")
def export_volunteer_excel(request):

    selected_user = request.GET.get("user")

    reports = LoginReport.objects.filter(
        user__adduser__role="volunteer"
    ).select_related("user", "user__adduser")

    if selected_user:
        reports = reports.filter(user_id=selected_user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Volunteer Report"

    ws.append([
        "Employee ID",
        "Name",
        "Date",
        "Login Time",
        "Logout Time"
    ])

    for r in reports:

        ws.append([
            r.user.adduser.employee_id,
            r.user.adduser.fullname,
            r.created_at.strftime("%d-%m-%Y"),
            r.login_time.strftime("%I:%M %p"),
            r.logout_time.strftime("%I:%M %p") if r.logout_time else ""
        ])

    response = HttpResponse(
        content_type="application/ms-excel"
    )

    response["Content-Disposition"] = "attachment; filename=volunteer_report.xlsx"

    wb.save(response)

    return response

@login_required(login_url="/")
def export_volunteer_pdf(request):

    selected_user = request.GET.get("user")

    reports = LoginReport.objects.filter(
        user__adduser__role="volunteer"
    ).select_related("user", "user__adduser")

    if selected_user:
        reports = reports.filter(user_id=selected_user)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=volunteer_report.pdf"

    p = canvas.Canvas(response)

    y = 800

    p.drawString(50, y, "Volunteer Report")
    y -= 40

    for r in reports:

        text = f"{r.user.adduser.employee_id} - {r.user.adduser.fullname} | Login: {r.login_time} | Logout: {r.logout_time}"

        p.drawString(50, y, text)

        y -= 20

        if y < 100:
            p.showPage()
            y = 800

    p.save()

    return response

@login_required(login_url="/")
def volunteeractivity(request):
    return render(request, 'volunteeractivity.html')

@login_required(login_url="/")
def viewfamily(request,id):

    family = get_object_or_404(Family, id=id)
    members = Member.objects.filter(family=family)
    all_families = Family.objects.exclude(id=id).only('id', 'family_id', 'head_name')

    return render(request, 'view-family.html', {
        "family": family, 
        "members": members,
        "all_families": all_families
    })

@login_required(login_url="/")
def remove_member_action(request, member_id):
    if request.method == "POST":
        member = get_object_or_404(Member, id=member_id)
        current_family = member.family
        reason = request.POST.get("reason")
        marriage_action = request.POST.get("marriage_action")
        target_family_id = request.POST.get("target_family")
        new_head_id = request.POST.get("new_head_id")

        # If the removed member is the family head, assign the new head
        if current_family and member.relationship and member.relationship.name == 'Self':
            if new_head_id:
                new_head = get_object_or_404(Member, id=new_head_id)
                try:
                    from .models import Relationship
                    self_rel = Relationship.objects.get(name='Self')
                    new_head.relationship = self_rel
                    new_head.save()
                except Exception:
                    pass
                current_family.head_name = new_head.name
                current_family.save()

        import datetime
        today = datetime.date.today()

        if reason == "Exited":
            member.family = None
            member.removal_reason = "Expired (Exited)"
            member.transfer_date = today
            member.save()
            messages.success(request, f"{member.name} has been marked as Exited.")
        elif reason == "Marriage":
            if marriage_action == "Remove":
                member.family = None
                member.removal_reason = "Marriage - Removed"
                member.transfer_date = today
                member.save()
                messages.success(request, f"{member.name} has been removed from the family due to marriage.")
            elif marriage_action == "Transfer" and target_family_id:
                target_family = get_object_or_404(Family, id=target_family_id)
                member.family = target_family
                # When transferring, clear the member_id so it follows the new family pattern.
                member.member_id = ""
                member.removal_reason = "Marriage - Transferred"
                member.transfer_date = today
                member.transferred_to_family = target_family
                member.save()
                messages.success(request, f"{member.name} has been transferred to family {target_family.family_id}.")
        
        return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
    return redirect('dashboard')

@login_required(login_url="/")
def removed_member_list(request):
    members = Member.objects.filter(family=None).order_by("-id")
    return render(request, "remove-memberlist.html", {"members": members})

@login_required(login_url="/")
def export_advance_excel(request):
    q = request.GET.get('q', '').strip()

    members = Member.objects.all().select_related(
        'family', 'blood_group', 'occupation', 'qualification',
        'income', 'overall_health', 'chronic_disease', 'relationship'
    )

    if q:
        members = members.filter(
            Q(name__icontains=q) |
            Q(member_id__icontains=q) |
            Q(family__family_id__icontains=q) |
            Q(alias__icontains=q) |
            Q(father_name__icontains=q) |
            Q(aadhaar_number__icontains=q) |
            Q(member_type__icontains=q) |
            Q(gender__icontains=q) |
            Q(marital_status__icontains=q) |
            Q(date_of_birth__icontains=q) |
            Q(relationship__name__icontains=q) |
            Q(mobile__icontains=q) |
            Q(whatsapp__icontains=q) |
            Q(email_id__icontains=q) |
            Q(blood_group__name__icontains=q) |
            Q(overall_health__name__icontains=q) |
            Q(chronic_disease__name__icontains=q) |
            Q(disability__icontains=q) |
            Q(activity_level__icontains=q) |
            Q(family__area__name__icontains=q) |
            Q(family__area__city__name__icontains=q) |
            Q(qualification__name__icontains=q) |
            Q(occupation__name__icontains=q) |
            Q(income__name__icontains=q) |
            Q(designation__icontains=q) |
            Q(department__icontains=q) |
            Q(company_name__icontains=q) |
            Q(skills__icontains=q) |
            Q(diploma_degree__icontains=q) |
            Q(sector__icontains=q) |
            Q(industry__icontains=q) |
            Q(membership_number__icontains=q) |
            Q(product_service_listing_in_yellow_pages__icontains=q) |
            Q(languages_speak__icontains=q) |
            Q(languages_read__icontains=q) |
            Q(languages_write__icontains=q) |
            Q(computer_proficiency__icontains=q) |
            Q(current_education_status__icontains=q) |
            Q(school_college_institute_name__icontains=q) |
            Q(grade_year__icontains=q) |
            Q(fees_payment__icontains=q) |
            Q(annual_academic_fees__icontains=q) |
            Q(sports__icontains=q) |
            Q(hobbies__icontains=q) |
            Q(career_goal__icontains=q) |
            Q(holy_koran_reading__icontains=q) |
            Q(deeniyath__icontains=q)
        ).distinct()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Advance Report"

    # Header row
    ws.append([
        "S.No", "Member ID", "Member Name", "Family ID",
        "Member Type", "Gender", "Blood Group", "Relationship",
        "Mobile", "Email", "Date of Birth", "Marital Status",
        "Qualification", "Occupation", "Annual Income",
        "Overall Health", "Chronic Disease", "Disability",
        "Area", "School / College", "Grade / Year",
        "Search Query"
    ])

    for idx, m in enumerate(members, start=1):
        ws.append([
            idx,
            m.member_id,
            m.name,
            m.family.family_id if m.family else "",
            m.member_type,
            m.gender or "",
            m.blood_group.name if m.blood_group else "",
            m.relationship.name if m.relationship else "",
            m.mobile or "",
            m.email_id or "",
            str(m.date_of_birth) if m.date_of_birth else "",
            m.marital_status or "",
            m.qualification.name if m.qualification else "",
            m.occupation.name if m.occupation else "",
            m.income.name if m.income else "",
            m.overall_health.name if m.overall_health else "",
            m.chronic_disease.name if m.chronic_disease else "",
            m.disability or "",
            m.family.area.name if m.family and m.family.area else "",
            m.school_college_institute_name or "",
            m.grade_year or "",
            q,
        ])

    response = HttpResponse(content_type="application/ms-excel")
    filename = f"advance_report_{q.replace(' ', '_') if q else 'all'}.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response

@login_required(login_url="/")
def export_advance_pdf(request):
    q = request.GET.get('q', '').strip()

    members = Member.objects.all().select_related(
        'family', 'blood_group', 'occupation', 'qualification',
        'income', 'overall_health', 'chronic_disease', 'relationship'
    )

    if q:
        members = members.filter(
            Q(name__icontains=q) |
            Q(member_id__icontains=q) |
            Q(family__family_id__icontains=q) |
            Q(alias__icontains=q) |
            Q(father_name__icontains=q) |
            Q(aadhaar_number__icontains=q) |
            Q(member_type__icontains=q) |
            Q(gender__icontains=q) |
            Q(marital_status__icontains=q) |
            Q(date_of_birth__icontains=q) |
            Q(relationship__name__icontains=q) |
            Q(mobile__icontains=q) |
            Q(whatsapp__icontains=q) |
            Q(email_id__icontains=q) |
            Q(blood_group__name__icontains=q) |
            Q(overall_health__name__icontains=q) |
            Q(chronic_disease__name__icontains=q) |
            Q(disability__icontains=q) |
            Q(activity_level__icontains=q) |
            Q(family__area__name__icontains=q) |
            Q(family__area__city__name__icontains=q) |
            Q(qualification__name__icontains=q) |
            Q(occupation__name__icontains=q) |
            Q(income__name__icontains=q) |
            Q(designation__icontains=q) |
            Q(department__icontains=q) |
            Q(company_name__icontains=q) |
            Q(skills__icontains=q) |
            Q(diploma_degree__icontains=q) |
            Q(sector__icontains=q) |
            Q(industry__icontains=q) |
            Q(membership_number__icontains=q) |
            Q(product_service_listing_in_yellow_pages__icontains=q) |
            Q(languages_speak__icontains=q) |
            Q(languages_read__icontains=q) |
            Q(languages_write__icontains=q) |
            Q(computer_proficiency__icontains=q) |
            Q(current_education_status__icontains=q) |
            Q(school_college_institute_name__icontains=q) |
            Q(grade_year__icontains=q) |
            Q(fees_payment__icontains=q) |
            Q(annual_academic_fees__icontains=q) |
            Q(sports__icontains=q) |
            Q(hobbies__icontains=q) |
            Q(career_goal__icontains=q) |
            Q(holy_koran_reading__icontains=q) |
            Q(deeniyath__icontains=q)
        ).distinct()

    import io
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buffer = io.BytesIO()
    filename = f"advance_report_{q.replace(' ', '_') if q else 'all'}.pdf"

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10*mm, leftMargin=10*mm,
        topMargin=12*mm, bottomMargin=12*mm
    )
    elements = []

    title_style = ParagraphStyle('title', fontSize=14, fontName='Helvetica-Bold',
                                 alignment=TA_CENTER, spaceAfter=4)
    sub_style   = ParagraphStyle('sub', fontSize=9, fontName='Helvetica',
                                 alignment=TA_CENTER, spaceAfter=10, textColor=colors.grey)

    elements.append(Paragraph("Advance Search Report", title_style))
    if q:
        elements.append(Paragraph(f'Search Query: "{q}"  |  Total Results: {members.count()}', sub_style))
    else:
        elements.append(Paragraph(f'All Members  |  Total: {members.count()}', sub_style))
    elements.append(Spacer(1, 4*mm))

    col_headers = ["#", "Member ID", "Name", "Family ID",
                   "Type", "Gender", "Blood Group", "Area",
                   "Mobile", "Qualification", "Occupation"]

    table_data = [col_headers]
    for idx, m in enumerate(members, start=1):
        table_data.append([
            str(idx),
            m.member_id or "",
            m.name or "",
            m.family.family_id if m.family else "",
            m.member_type or "",
            m.gender or "",
            m.blood_group.name if m.blood_group else "",
            m.family.area.name if m.family and m.family.area else "",
            m.mobile or "",
            m.qualification.name if m.qualification else "",
            m.occupation.name if m.occupation else "",
        ])

    col_widths = [8*mm, 28*mm, 40*mm, 24*mm,
                  18*mm, 16*mm, 22*mm, 28*mm,
                  26*mm, 30*mm, 28*mm]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#4361ee')),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0), 8),
        ('ALIGN',         (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING',    (0, 0), (-1, 0), 6),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 7.5),
        ('ALIGN',         (0, 1), (0, -1), 'CENTER'),
        ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9ff')]),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#e0e6ef')),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]))
    elements.append(t)

    doc.build(elements)
    pdf_data = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@login_required(login_url="/")
def volunteer_familylist(request):
    families = Family.objects.all()
    return render(request, 'volunteer-familylist.html', {"families": families})


@login_required(login_url="/")
def volunteer_viewfamily(request, id):
    family = Family.objects.get(id=id)
    members = Member.objects.filter(family=family)
    return render(request, 'volunteer-viewfamily.html', {"family": family, "members": members})